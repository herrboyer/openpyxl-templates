from datetime import datetime
from itertools import chain
from os.path import dirname, join

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font

from columns import IntegerColumn, ChoiceColumn, DateTimeColumn, TimeColumn, CharColumn
from style import CellStyle
from utils import SolidFill
from workbook import WorkbookTemplate
from worksheet import SheetTemplate, RowExceptionPolicy

BASE_DIR = dirname(__file__)

choices = (
    ("lägg till", "ADD"),
    ("ta bort", "REMOVE"),
    ("inget", "NOOP"),
    (None, None),
)


class TestSheet(SheetTemplate):
    name = "TestSheet"
    title = "Test blad"
    header_style = CellStyle(font=Font(bold=True, color="FFFFFFFF"), fill=SolidFill("5d1738"))
    row_exception_policy = RowExceptionPolicy.RETURN_EXCEPTION

    columns = [
        CharColumn(
            object_attr="one",
            header="Röd text",
            width=15,
            style=CellStyle(
                font=Font(color="FFFF0000")
            )
        ),
        IntegerColumn(
            object_attr="three",
            header="Svart siffra",
            width=15
        ),
        IntegerColumn(
            object_attr="six",
            header="Ytterligare en siffra",
            width=15
        ),
        # EmptyColumn(
        #     header="Empty",
        #     header_style=CellStyle(fill=SolidFill("AAAAAA")),
        #     hidden=False,
        #     width=10
        # ),
        ChoiceColumn(
            object_attr="action",
            header="Aktion",
            choices=choices,
            width=30,
            default_value=""
            # default_value="ADD"
        ),
        DateTimeColumn(
            object_attr="date",
            header="Datum",
            width=15,
        ),
        TimeColumn(
            object_attr="time",
            header="Tid",
            width=10,
        )
    ]


class TestWorkbook(WorkbookTemplate):
    sheets = [
        TestSheet()
    ]
    active_sheet = "TestSheet"


test = TestWorkbook(load_workbook(join(BASE_DIR, "test1.xlsx").replace('\\', '/'), ))

for row in test.read_rows("TestSheet"):
    print(row)


class TestObject:
    def __init__(self, one, three, six, action, date):
        self.one = one
        self.three = three
        self.six = six

        self.action = action
        self.date = date

        self.time = date.time()


output_workbook = Workbook()
test_output = TestWorkbook(output_workbook, style=CellStyle(Font(name="Comic Sans MS")))
test_output.write_sheet("TestSheet", (
    TestObject(1, 2, 3, "ADD", datetime.now()),
    TestObject(2, 3, 4, "NOOP", datetime.now()),
    TestObject(3, 4, 5, "REMOVE", datetime.now()),
    TestObject(4, 5, 6, "ADD", datetime.now()),
))

output_workbook.save(join(BASE_DIR, "test_output.xlsx").replace('\\', '/'))

alignment = Alignment(wrap_text=False)
border = Border()
font = Font(size=24)

for x in chain(alignment, font):
    print(x)